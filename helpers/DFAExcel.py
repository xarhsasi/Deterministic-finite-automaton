from openpyxl import Workbook


class DFAExcel:
    """
        DFAExcel class which creates an excel file with the dfa table for the automato
    """

    def __init__(self, filename="dfa_sample.xlsx"):
        """
        Initialize the DFAExcel instance
        :param filename:
        """
        self.wb = Workbook()
        self.ws = None
        self.filename = filename
        self.activate_wb()

    def activate_wb(self):
        """
        Activate the workbook
        :return: void
        """
        self.ws = self.wb.active
        self.initialize_wb()

    def initialize_wb(self):
        """
        Initialize the workbook
        :return: void
        """
        self.ws.title = "DFA Workbook"

    def write_to_excel(self, table, prefix):
        """
        File logic and format
        :param table:
        :param prefix:
        :return:
        """
        self.ws['A1'] = "node"
        self.ws['B1'] = "state"
        self.ws['C1'] = "next node"
        self.ws['D1'] = "is first state"
        self.ws['E1'] = "is final state"

        row_i = 2
        for row in table.rows:
            self.ws.cell(column=1, row=row_i, value=prefix + row.current_state)
            self.ws.cell(column=2, row=row_i, value=row.state)
            self.ws.cell(column=3, row=row_i, value=prefix + row.next_state)
            self.ws.cell(column=4, row=row_i, value=row.is_first)
            self.ws.cell(column=5, row=row_i, value=row.is_final)
            row_i = row_i + 1
        self.save_wb()

    def save_wb(self):
        """
        Save the workbook
        :return:
        """
        self.wb.save(self.filename)
